import streamlit as st
import os
import openpyxl

# Function for handling temporary file creation and management
def handle_temp_file(uploaded_file, temp_dir="./temp"):
    temp_file_path = os.path.join(temp_dir, uploaded_file.name)
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    try:
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return temp_file_path
    except Exception as e:
        st.error(f"Error handling temporary file: {e}")
        return None

# Function to delete specific sheets from the workbook
def delete_sheets(workbook, sheet_names):
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

# Function to clear rows based on license selection
def clear_rows(sheet, start_row, end_row):
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.value = ""

# Function to remove hyperlinks at specific cells
def remove_hyperlinks(sheet, cells):
    for cell in cells:
        if sheet[cell].hyperlink is not None:
            sheet[cell].hyperlink = None

# File handling and editing subpart
def file_editing():
    st.subheader("File Editing")

    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"], key="file_uploader")

    if uploaded_file is not None:
        temp_file_path = handle_temp_file(uploaded_file)
        if temp_file_path is None:
            st.stop()
            return

        try:
            workbook = openpyxl.load_workbook(temp_file_path)

            sheets_to_delete = [
                "Change Log",
                "Info",
                "Threat Policy (Endpoint)",
                "Threat Policy (Server)",
                "Exceptions"
            ]
            
            delete_sheets(workbook, sheets_to_delete)

            if "Global Settings" in workbook.sheetnames:
                global_settings_sheet = workbook["Global Settings"]
                global_settings_sheet.title = "General Settings"
                global_settings_sheet["A1"] = "General Settings"

                if st.sidebar.checkbox("Endpoint + Server Complete"):
                    delete_sheets(workbook, sheets_to_delete)
                    global_settings_sheet["A6"] = "Live Response - Endpoint"
                    global_settings_sheet["A7"] = "Live Response - Server"
                if st.sidebar.checkbox("Endpoint + Server"):
                    delete_sheets(workbook, sheets_to_delete)
                    clear_rows(global_settings_sheet, 6, 7)
                    remove_hyperlinks(global_settings_sheet, ["A6", "A7"])
                if st.sidebar.checkbox("Endpoint License"):
                    clear_rows(global_settings_sheet, 6, 7)
                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "server" in sheet_name.lower()])
                if st.sidebar.checkbox("Server License"):
                    clear_rows(global_settings_sheet, 6, 7)
                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "endpoint" in sheet_name.lower()])
                if st.sidebar.checkbox("Endpoint License - Complete"):
                    global_settings_sheet["A6"] = "Live Response - Endpoint"
                    clear_rows(global_settings_sheet, 7, 7)
                    remove_hyperlinks(global_settings_sheet, ["A6", "A7"])

                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "server" in sheet_name.lower()])

                if st.sidebar.checkbox("Server License - Complete"):
                    global_settings_sheet["A6"] = "Live Response - Server"
                    clear_rows(global_settings_sheet, 7, 7)
                    remove_hyperlinks(global_settings_sheet, ["A6", "A7"])

                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "endpoint" in sheet_name.lower()])

            if "Calculations" in workbook.sheetnames:
                del workbook["Calculations"]

            edited_file_path = f"./temp/edited_{uploaded_file.name}"
            workbook.save(edited_file_path)

            try:
                with open(edited_file_path, 'rb') as f:
                    file_data = f.read()
                st.download_button(label='Click here to download the edited file',
                                   data=file_data,
                                   file_name=f"edited_{uploaded_file.name}",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key='file_download_button',
                                   help='Right-click and save as...')
                os.remove(edited_file_path)
            except Exception as e:
                st.error(f'An error occurred while downloading the file: {e}')

        except Exception as e:
            st.error(f"Error reading or editing Excel file: {e}")
            os.remove(temp_file_path)

# Start the file editing process
file_editing()
import streamlit as st
import os
import openpyxl

# Function for handling temporary file creation and management
def handle_temp_file(uploaded_file, temp_dir="./temp"):
    temp_file_path = os.path.join(temp_dir, uploaded_file.name)
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    try:
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return temp_file_path
    except Exception as e:
        st.error(f"Error handling temporary file: {e}")
        return None

# Function to delete specific sheets from the workbook
def delete_sheets(workbook, sheet_names):
    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

# Function to clear rows based on license selection
def clear_rows(sheet, start_row, end_row):
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            cell.value = ""

# Function to remove hyperlinks at specific cells
def remove_hyperlinks(sheet, cells):
    for cell in cells:
        if sheet[cell].hyperlink is not None:
            sheet[cell].hyperlink = None

# File handling and editing subpart
def file_editing():
    st.subheader("File Editing")

    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"], key="file_uploader")

    if uploaded_file is not None:
        temp_file_path = handle_temp_file(uploaded_file)
        if temp_file_path is None:
            st.stop()
            return

        try:
            workbook = openpyxl.load_workbook(temp_file_path)

            sheets_to_delete = [
                "Change Log",
                "Info",
                "Threat Policy (Endpoint)",
                "Threat Policy (Server)",
                "Exceptions"
            ]
            
            delete_sheets(workbook, sheets_to_delete)

            if "Global Settings" in workbook.sheetnames:
                global_settings_sheet = workbook["Global Settings"]
                global_settings_sheet.title = "General Settings"
                global_settings_sheet["A1"] = "General Settings"

                if st.sidebar.checkbox("All the Below Complete"):
                    delete_sheets(workbook, sheets_to_delete)
                    global_settings_sheet["A6"] = "Live Response - Endpoint"
                    global_settings_sheet["A7"] = "Live Response - Server"
                if st.sidebar.checkbox("All the Below Normal"):
                    delete_sheets(workbook, sheets_to_delete)
                    clear_rows(global_settings_sheet, 6, 7)
                    remove_hyperlinks(global_settings_sheet, ["A6", "A7"])
                if st.sidebar.checkbox("Endpoint License"):
                    clear_rows(global_settings_sheet, 6, 7)
                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "server" in sheet_name.lower()])
                if st.sidebar.checkbox("Server License"):
                    clear_rows(global_settings_sheet, 6, 7)
                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "endpoint" in sheet_name.lower()])
                if st.sidebar.checkbox("Endpoint License - Complete"):
                    global_settings_sheet["A6"] = "Live Response - Endpoint"
                    clear_rows(global_settings_sheet, 7, 7)
                    remove_hyperlinks(global_settings_sheet, ["A6", "A7"])

                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "server" in sheet_name.lower()])

                if st.sidebar.checkbox("Server License - Complete"):
                    global_settings_sheet["A6"] = "Live Response - Server"
                    clear_rows(global_settings_sheet, 7, 7)
                    remove_hyperlinks(global_settings_sheet, ["A6", "A7"])

                    delete_sheets(workbook, [sheet_name for sheet_name in workbook.sheetnames if "endpoint" in sheet_name.lower()])

            if "Calculations" in workbook.sheetnames:
                del workbook["Calculations"]

            edited_file_path = f"./temp/edited_{uploaded_file.name}"
            workbook.save(edited_file_path)

            try:
                with open(edited_file_path, 'rb') as f:
                    file_data = f.read()
                st.download_button(label='Click here to download the edited file',
                                   data=file_data,
                                   file_name=f"edited_{uploaded_file.name}",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key='file_download_button',
                                   help='Right-click and save as...')
                os.remove(edited_file_path)
            except Exception as e:
                st.error(f'An error occurred while downloading the file: {e}')

        except Exception as e:
            st.error(f"Error reading or editing Excel file: {e}")
            os.remove(temp_file_path)

# Start the file editing process
file_editing()
